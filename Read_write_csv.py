import os
from os.path import exists
import pandas as pd
import csv
from openpyxl import Workbook
from openpyxl import load_workbook


def print_mortgage_results(m):
    # Open a new CSV file in write mode
    mortgage_results = '/Users/michaelaneiro/Downloads/mortgage_results_out.csv'  # MacBook

    # Delete the previous output file
    if exists(mortgage_results):
        os.remove(mortgage_results)

    with open(mortgage_results, 'w', newline='') as csvfile:
        # Create a CSV writer object
        csvwriter = csv.writer(csvfile)
        # Write the header row
        csvwriter.writerow(['Balance', 'Scheduled Pay', 'Scheduled Prin', 'Scheduled Interest', 'Prepay', 'Default', 'Loss'])
        # Write each MortgageCashflows object to a row in the CSV file
        for i in range(m.amortization_term):
            csvwriter.writerow([m.cashflows[i].balance, m.cashflows[i].scheduled_payment, m.cashflows[i].scheduled_principal,  m.cashflows[i].scheduled_interest, m.cashflows[i].prepayment_principal, m.cashflows[i].default_principal, m.cashflows[i].default_loss])

    return


def print_mortgage_results_v2(m, r):  # pass the Mortgage object and the iteration of loan analysis
    # Open a new XLSX file in write mode
    mortgage_results = '/Users/michaelaneiro/Downloads/mortgage_results_out.xlsx'

    # Create a new DataFrame to store the results
    df = pd.DataFrame(columns=['Balance', 'Scheduled Pay', 'Scheduled Prin', 'Scheduled Interest', 'Prepay', 'Default', 'Loss'])

    for i in range(m.amortization_term):
        df.loc[i] = [m.cashflows[i].balance, m.cashflows[i].scheduled_payment, m.cashflows[i].scheduled_principal,  m.cashflows[i].scheduled_interest, m.cashflows[i].prepayment_principal, m.cashflows[i].default_principal, m.cashflows[i].default_loss]

    if exists(mortgage_results):  # temporary cheat until I correct lower code logic
        os.remove(mortgage_results)

    # Check if the file exists
    if not exists(mortgage_results):
        # If the file doesn't exist, create a new Excel file with the specific sheet name
        sheet_name = "Mortgage_Results0"
        with pd.ExcelWriter(mortgage_results, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        # If the file exists, update the existing Excel file with the new data
        sheet_name = "Mortgage_Results" + str(r)
        book = None
        try:
            book = pd.read_excel(mortgage_results, engine='openpyxl', sheet_name=None)
        except Exception as e:
            print(f"Error: {e}")
            return

        workbook = load_workbook(mortgage_results)

        # Check if the sheet exists
        with pd.ExcelWriter(mortgage_results, engine='openpyxl', mode='a') as writer:
            if sheet_name not in book:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                if sheet_name in workbook.sheetnames:
                    # Delete the sheet
                    workbook.remove(workbook[sheet_name])
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    workbook = load_workbook(mortgage_results)

    return
